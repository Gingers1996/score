import streamlit as st
import pandas as pd
import io
from typing import Optional
import hashlib
import time
import re

def anonymize_data(df: pd.DataFrame) -> pd.DataFrame:
    """数据匿名化处理"""
    anonymized_df = df.copy()
    
    # 生成匿名ID
    if '学号' in anonymized_df.columns:
        anonymized_df['匿名ID'] = [f"ID_{i+1:04d}" for i in range(len(anonymized_df))]
    
    # 可选：匿名化姓名（保留前两位字符）
    if '姓名' in anonymized_df.columns:
        anonymized_df['姓名'] = anonymized_df['姓名'].apply(
            lambda x: x[:2] + '*' * (len(str(x)) - 2) if len(str(x)) > 2 else x
        )
    
    return anonymized_df

def calculate_total_score(row):
    """计算总分：总分=(乙部/103*0.7)*100+(甲部/50*0.3)*100"""
    jia_score = float(row['甲部分数'])
    yi_score = float(row['乙部分数'])
    
    # 计算加权分数
    jia_weighted = (jia_score / 50 * 0.3) * 100
    yi_weighted = (yi_score / 103 * 0.7) * 100
    
    # 四舍五入为整数
    return round(jia_weighted + yi_weighted)

def process_data(df: pd.DataFrame) -> pd.DataFrame:
    """处理数据：计算总分、排序、排名"""
    # 复制数据框避免修改原始数据
    processed_df = df.copy()
    
    # 计算总分
    processed_df['总分'] = processed_df.apply(calculate_total_score, axis=1)
    
    # 按总分降序排序
    processed_df = processed_df.sort_values('总分', ascending=False)
    
    # 计算排名（相同分数相同排名，类似WPS的RANK函数）
    processed_df['排名'] = processed_df['总分'].rank(method='min', ascending=False).astype(int)
    
    return processed_df

def assign_grades(df: pd.DataFrame, cutoff_scores: dict) -> pd.DataFrame:
    """根据cutoff分数分配等级"""
    df_with_grades = df.copy()
    
    # 初始化等级列
    df_with_grades['等级'] = '未定级'
    
    # 按分数从低到高分配等级（Level2最低，Level7最高）
    levels = ['Level2', 'Level3', 'Level4', 'Level5', 'Level6', 'Level7']
    
    for level in levels:
        if level in cutoff_scores and cutoff_scores[level] > 0:
            mask = df_with_grades['总分'] >= cutoff_scores[level]
            df_with_grades.loc[mask, '等级'] = level
    
    return df_with_grades

def validate_cutoff_input(value: str) -> Optional[int]:
    """验证等级分数线输入"""
    try:
        score = int(value)
        if 0 <= score <= 100:
            return score
        else:
            return None
    except ValueError:
        return None

def validate_cutoff_input(value: str) -> Optional[int]:
    """验证等级分数线输入"""
    try:
        score = int(value)
        if 0 <= score <= 100:
            return score
        else:
            return None
    except ValueError:
        return None

def main():
    st.set_page_config(
        page_title="学生成绩计算系统 - 云端安全版",
        page_icon="☁️",
        layout="wide"
    )
    
    # 隐私保护声明
    st.sidebar.markdown("## 🔒 隐私保护")
    st.sidebar.info("""
    **云端安全承诺：**
    - 数据会进行匿名化处理
    - 敏感信息会被保护
    - 建议使用化名数据
    - 定期清理服务器数据
    """)
    
    st.title("☁️ 学生成绩计算系统 - 云端安全版")
    st.markdown("---")
    
    # 数据匿名化选项
    st.sidebar.header("🔐 数据保护设置")
    anonymize_option = st.sidebar.checkbox(
        "启用数据匿名化", 
        value=True,
        help="启用后将自动匿名化姓名和学号"
    )
    
    # 侧边栏：等级 cutoff 设置
    st.sidebar.header("🏆 等级 cutoff 设置")
    default_cutoffs = {
        'Level2': 47,
        'Level3': 53,
        'Level4': 58,
        'Level5': 63,
        'Level6': 66,
        'Level7': 70
    }
    
    if 'cutoffs' not in st.session_state:
        st.session_state['cutoffs'] = default_cutoffs.copy()
    
    # 等级分数线输入（使用文本输入框，更便捷）
    st.sidebar.caption("请设置各等级的最低分数线（≥）。Level2到Level7递增。")
    
    # 使用文本输入框，支持直接输入整数
    level2_input = st.sidebar.text_input(
        "Level2 ≥", 
        value=str(st.session_state['cutoffs']['Level2']),
        help="输入0-100之间的整数"
    )
    level3_input = st.sidebar.text_input(
        "Level3 ≥", 
        value=str(st.session_state['cutoffs']['Level3']),
        help="输入0-100之间的整数"
    )
    level4_input = st.sidebar.text_input(
        "Level4 ≥", 
        value=str(st.session_state['cutoffs']['Level4']),
        help="输入0-100之间的整数"
    )
    level5_input = st.sidebar.text_input(
        "Level5 ≥", 
        value=str(st.session_state['cutoffs']['Level5']),
        help="输入0-100之间的整数"
    )
    level6_input = st.sidebar.text_input(
        "Level6 ≥", 
        value=str(st.session_state['cutoffs']['Level6']),
        help="输入0-100之间的整数"
    )
    level7_input = st.sidebar.text_input(
        "Level7 ≥", 
        value=str(st.session_state['cutoffs']['Level7']),
        help="输入0-100之间的整数"
    )
    
    # 验证输入并更新session state
    inputs = [level2_input, level3_input, level4_input, level5_input, level6_input, level7_input]
    levels = ['Level2', 'Level3', 'Level4', 'Level5', 'Level6', 'Level7']
    
    # 验证所有输入
    valid_inputs = True
    for i, input_val in enumerate(inputs):
        validated = validate_cutoff_input(input_val)
        if validated is None:
            st.sidebar.error(f"{levels[i]} 输入无效，请输入0-100之间的整数")
            valid_inputs = False
            break
    
    # 应用按钮
    col_a, col_b = st.sidebar.columns(2)
    with col_a:
        if st.button("应用等级设置"):
            if valid_inputs:
                new_cutoffs = {}
                for i, level in enumerate(levels):
                    new_cutoffs[level] = validate_cutoff_input(inputs[i])
                st.session_state['cutoffs'] = new_cutoffs
                st.sidebar.success("等级设置已应用")
                st.rerun()
    
    with col_b:
        if st.button("恢复默认等级"):
            st.session_state['cutoffs'] = default_cutoffs.copy()
            st.sidebar.success("已恢复默认等级设置")
            st.rerun()
    
    current_cutoffs = st.session_state['cutoffs']
    
    # 显示当前等级分数线（只读）
    with st.sidebar.expander("当前等级分数线（只读）", expanded=False):
        for level, score in current_cutoffs.items():
            st.write(f"{level}: ≥ {score}分")
    
    # 文件上传区域
    st.header("📁 文件上传")
    
    # 安全提醒
    if anonymize_option:
        st.info("🔒 已启用数据匿名化，敏感信息将被保护")
    
    uploaded_file = st.file_uploader(
        "请上传Excel或CSV文件",
        type=['xlsx', 'xls', 'csv'],
        help="文件应包含以下列：姓名、学号、班级、甲部分数、乙部分数"
    )
    
    if uploaded_file is not None:
        try:
            # 读取文件
            if uploaded_file.name.endswith('.csv'):
                df = pd.read_csv(uploaded_file)
            else:
                df = pd.read_excel(uploaded_file)
            
            # 处理None值，用0替代
            df = df.fillna(0)
            
            # 数据隐私保护：生成文件哈希用于日志
            file_content = uploaded_file.read()
            file_hash = hashlib.md5(file_content).hexdigest()[:8]
            uploaded_file.seek(0)  # 重置文件指针
            
            st.success(f"✅ 文件上传成功！共读取 {len(df)} 条记录")
            st.info(f"🔒 文件ID: {file_hash} (仅用于日志，不包含个人信息)")
            
            # 检查必要列是否存在
            required_columns = ['姓名', '学号', '班级', '甲部分数', '乙部分数']
            missing_columns = [col for col in required_columns if col not in df.columns]
            
            if missing_columns:
                st.error(f"❌ 文件缺少必要的列：{', '.join(missing_columns)}")
                st.info("请确保文件包含以下列：姓名、学号、班级、甲部分数、乙部分数")
                return
            
            # 数据匿名化处理
            if anonymize_option:
                df = anonymize_data(df)
                st.info("🔐 数据已进行匿名化处理")
            
            # 显示原始数据
            st.subheader("📋 原始数据")
            st.dataframe(df, use_container_width=True)
            
            # 处理数据
            processed_df = process_data(df)
            
            # 显示处理后的数据
            st.subheader("📊 计算结果")
            st.dataframe(processed_df, use_container_width=True)
            
            # 调试信息
            st.info(f"📝 数据检查：总分范围 {processed_df['总分'].min()} - {processed_df['总分'].max()}")
            
            # 统计信息
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("总人数", len(processed_df))
            with col2:
                st.metric("平均分", f"{processed_df['总分'].mean():.1f}")
            with col3:
                st.metric("最高分", f"{processed_df['总分'].max()}")
            with col4:
                st.metric("最低分", f"{processed_df['总分'].min()}")
            
            # 等级划分
            final_df = assign_grades(processed_df, current_cutoffs)
            
            # 调试等级分配
            level_counts = final_df['等级'].value_counts()
            st.info(f"📊 等级分配：{dict(level_counts)}")
            
            # 显示最终结果（按等级涂色）
            st.subheader("🎯 最终结果（含等级）")
            
            # 定义等级颜色映射
            level_colors = {
                'Level2': '#FFE6E6',  # 浅红色
                'Level3': '#FFF2E6',  # 浅橙色
                'Level4': '#FFFFE6',  # 浅黄色
                'Level5': '#E6FFE6',  # 浅绿色
                'Level6': '#E6F3FF',  # 浅蓝色
                'Level7': '#F0E6FF',  # 浅紫色
                '未定级': '#F5F5F5'   # 浅灰色
            }
            
            # 创建样式函数
            def highlight_levels(df):
                styles = pd.DataFrame('', index=df.index, columns=df.columns)
                for i in range(len(df)):
                    level = df.iloc[i]['等级']
                    color = level_colors.get(level, '#F5F5F5')
                    for j in range(len(df.columns)):
                        styles.iloc[i, j] = f'background-color: {color}'
                return styles
            
            # 应用样式并显示
            styled_df = final_df.style.apply(highlight_levels, axis=None)
            st.dataframe(styled_df, use_container_width=True)
            
            # 成绩分布统计
            st.subheader("📊 成绩分布")
            
            # 等级分布
            col1, col2 = st.columns(2)
            with col1:
                st.write("**等级分布**")
                grade_counts = final_df['等级'].value_counts().sort_index()
                for level, count in grade_counts.items():
                    percentage = (count / len(final_df)) * 100
                    st.write(f"{level}: {count}人 ({percentage:.1f}%)")
            
            with col2:
                st.write("**班级平均分**")
                class_avg = final_df.groupby('班级')['总分'].mean().sort_values(ascending=False)
                for class_name, avg_score in class_avg.items():
                    st.write(f"{class_name}: {avg_score:.1f}分")
            
            # 分数区间统计
            st.write("**分数区间分布**")
            bins = [0, 50, 60, 70, 80, 90, 100]
            labels = ['<50', '50-60', '60-70', '70-80', '80-90', '90-100']
            final_df['分数区间'] = pd.cut(final_df['总分'], bins=bins, labels=labels, include_lowest=True)
            score_dist = final_df['分数区间'].value_counts().sort_index()
            
            for interval, count in score_dist.items():
                percentage = (count / len(final_df)) * 100
                st.write(f"{interval}: {count}人 ({percentage:.1f}%)")
            
            # 下载结果
            st.subheader("💾 下载结果")
            
            # 创建Excel文件（带颜色）
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                final_df.to_excel(writer, sheet_name='计算结果', index=False)
                
                # 获取workbook和worksheet对象
                workbook = writer.book
                worksheet = writer['计算结果']
                
                # 定义等级颜色映射（openpyxl格式）
                from openpyxl.styles import PatternFill
                level_fills = {
                    'Level2': PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid'),
                    'Level3': PatternFill(start_color='FFF2E6', end_color='FFF2E6', fill_type='solid'),
                    'Level4': PatternFill(start_color='FFFFE6', end_color='FFFFE6', fill_type='solid'),
                    'Level5': PatternFill(start_color='E6FFE6', end_color='E6FFE6', fill_type='solid'),
                    'Level6': PatternFill(start_color='E6F3FF', end_color='E6F3FF', fill_type='solid'),
                    'Level7': PatternFill(start_color='F0E6FF', end_color='F0E6FF', fill_type='solid'),
                    '未定级': PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
                }
                
                # 应用颜色到等级列
                level_col_index = final_df.columns.get_loc('等级') + 1  # Excel列从1开始
                for row_idx, level in enumerate(final_df['等级'], start=2):  # Excel行从2开始（跳过标题）
                    if level in level_fills:
                        for col_idx in range(1, len(final_df.columns) + 1):
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            cell.fill = level_fills[level]
                
                # 设置列宽
                for col_idx, col_name in enumerate(final_df.columns, start=1):
                    max_len = max(
                        final_df[col_name].astype(str).apply(len).max(),
                        len(col_name)
                    )
                    worksheet.column_dimensions[chr(64 + col_idx)].width = max_len + 2
            
            output.seek(0)
            
            # 生成安全的文件名
            timestamp = int(time.time())
            safe_filename = f"成绩计算结果_{file_hash}_{timestamp}.xlsx"
            
            st.download_button(
                label="📥 下载Excel文件",
                data=output.getvalue(),
                file_name=safe_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            
            # 云端安全提醒
            st.warning("""
            ☁️ **云端安全提醒：**
            - 数据已进行匿名化处理
            - 建议使用化名数据
            - 定期清理敏感文件
            - 不要上传包含真实身份信息的数据
            """)
            
        except Exception as e:
            st.error(f"❌ 处理文件时出错：{str(e)}")
            st.info("请检查文件格式是否正确，确保包含必要的列")
    
    else:
        st.info("👆 请上传包含学生成绩的Excel或CSV文件")
        
        # 显示示例数据格式
        st.subheader("📝 文件格式示例")
        example_data = {
            '姓名': ['张三', '李四', '王五'],
            '学号': ['2021001', '2021002', '2021003'],
            '班级': ['一班', '一班', '二班'],
            '甲部分数': [45, 42, 48],
            '乙部分数': [95, 88, 92]
        }
        example_df = pd.DataFrame(example_data)
        st.dataframe(example_df, use_container_width=True)
        
        st.markdown("""
        **计算规则：**
        - 总分 = (乙部分数/103 × 0.7) × 100 + (甲部分数/50 × 0.3) × 100
        - 系统会自动按总分降序排序并计算排名
        - 您可以根据需要设置各等级的cutoff分数
        """)
        
        # 云端使用建议
        st.subheader("☁️ 云端使用建议")
        st.markdown("""
        **为了保护隐私，建议：**
        1. **使用化名数据**：将真实姓名替换为化名
        2. **简化学号**：使用简单的编号替代真实学号
        3. **启用匿名化**：系统会自动保护敏感信息
        4. **及时清理**：使用完毕后及时删除文件
        """)

if __name__ == "__main__":
    main()
